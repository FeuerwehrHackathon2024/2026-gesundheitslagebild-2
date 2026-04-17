"""Verteilungs-Engine + XLSX-Parser + Belegungs-Simulation."""

from __future__ import annotations

import io
import logging
import math
import random
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy import and_, or_

from .extensions import db
from .here_client import fetch_route
from .ivena_mapping import SK_TRANSPORTMITTEL as IVENA_TRANSPORTMITTEL
from .models import (
    Fahrt,
    Hub,
    Krankenhaus,
    KrankenhausBelegung,
    Patient,
    PatientenBatch,
    TransportAuftrag,
)

log = logging.getLogger(__name__)

# SK → Aufenthaltsdauer Tage
SK_AUFENTHALT_TAGE = {"SK1": 5, "SK2": 3, "SK3": 1}

# SK → empfohlenes Transportmittel
SK_TRANSPORTMITTEL = {"SK1": "RTW", "SK2": "KTW", "SK3": "Taxi"}

# Pickup-Offset je Transportmittel in Minuten (bis der Patient am Hub abgeholt ist)
TRANSPORTMITTEL_PICKUP_OFFSET_MIN = {"RTW": 2, "KTW": 5, "BTW": 5, "Taxi": 10}

# Fahrzeug-Kapazitäten (Anzahl Patienten pro Fahrt) — gemäß Vorgabe
TRANSPORTMITTEL_KAPAZITAET = {
    "RTW": 1,   # Rettungswagen mit Notarzt: 1 Patient (vital)
    "KTW": 1,   # Krankentransportwagen: 1 Patient (liegend)
    "BTW": 2,   # Behindertentransportwagen: 2 Patienten (sitzend/Rollstuhl)
    "Taxi": 2,  # Taxi/Patiententransport: 2 Patienten (sitzend, mobil)
}

# SK-Kompatibilität: ein SK1-fähiges KH kann auch SK2 und SK3 versorgen
SK_KANN_COLS = {
    "SK1": ["kann_sk1"],
    "SK2": ["kann_sk1", "kann_sk2"],
    "SK3": ["kann_sk1", "kann_sk2", "kann_sk3"],
}

SK_CAP_COLS = {"SK1": "kapazitaet_sk1", "SK2": "kapazitaet_sk2", "SK3": "kapazitaet_sk3"}


# ---------------- Helpers ----------------

def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6371.0
    from math import radians, sin, cos, asin, sqrt
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    return 2 * R * asin(sqrt(a))


# ---------------- XLSX Parse ----------------

def _normalize_sk(value) -> str | None:
    """Akzeptiert sowohl SK1/SK2/SK3 als auch IVENA-Notation (I, II, Rot, Gelb, …)."""
    from .ivena_mapping import map_ivena_to_sk
    return map_ivena_to_sk(value)


def _as_datetime(day: date | None, t: time | None) -> datetime | None:
    if day is None and t is None:
        return None
    if day is None:
        day = date.today()
    if t is None:
        t = time(0, 0, 0)
    if isinstance(t, datetime):
        return t
    return datetime.combine(day, t)


def parse_ivena_xlsx(raw: bytes, default_hub_name: str = "Hub Süd") -> list[dict]:
    """Liest IVENA-artige XLSX und liefert eine Liste Patienten-Dicts.

    Erwartete Spalten (flexibel): ID, SK, Datum,
        Eingangssichtung um, Transportbereitschaft gesetzt um, Quelle (optional)
    """
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Header normalisieren
    header_row = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def find(*names: str) -> int | None:
        for n in names:
            n_low = n.lower()
            for i, h in enumerate(header_row):
                if h == n_low or n_low in h:
                    return i
        return None

    col_id = find("id")
    col_sk = find("sk")
    col_datum = find("datum", "date")
    col_sichtung = find("eingangssichtung", "sichtung")
    col_transport = find("transportbereit", "transport")
    col_quelle = find("quelle", "hub", "origin", "herkunft")

    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        sk = _normalize_sk(row[col_sk]) if col_sk is not None else None
        if not sk:
            continue

        ext_id = str(row[col_id]).strip() if col_id is not None and row[col_id] else None
        datum_val = row[col_datum] if col_datum is not None else None
        sichtung_val = row[col_sichtung] if col_sichtung is not None else None
        transport_val = row[col_transport] if col_transport is not None else None
        quelle_val = row[col_quelle] if col_quelle is not None else None

        datum: date | None = None
        if isinstance(datum_val, datetime):
            datum = datum_val.date()
        elif isinstance(datum_val, date):
            datum = datum_val

        sichtung_dt = _as_datetime(datum, sichtung_val)
        transport_dt = _as_datetime(datum, transport_val)

        out.append({
            "external_id": ext_id,
            "sk": sk,
            "datum": datum,
            "eingangssichtung": sichtung_dt,
            "transportbereit": transport_dt,
            "quelle": (str(quelle_val).strip() if quelle_val else default_hub_name),
        })
    return out


# ---------------- Belegungs-Simulation ----------------

def init_belegung_rows() -> int:
    """Legt für jede Klinik mit Geo einen Belegungs-Eintrag an (idempotent)."""
    kh_rows = (
        db.session.query(Krankenhaus)
        .filter(Krankenhaus.lat.isnot(None))
        .all()
    )
    existing_ids = {
        row[0] for row in db.session.query(KrankenhausBelegung.krankenhaus_id).all()
    }
    created = 0
    for kh in kh_rows:
        if kh.id in existing_ids:
            continue
        row = KrankenhausBelegung(
            krankenhaus_id=kh.id,
            kapazitaet_sk1=kh.kapazitaet_sk1_geschaetzt or 0,
            kapazitaet_sk2=kh.kapazitaet_sk2_geschaetzt or 0,
            kapazitaet_sk3=kh.kapazitaet_sk3_geschaetzt or 0,
            belegung_sk1=0,
            belegung_sk2=0,
            belegung_sk3=0,
        )
        db.session.add(row)
        created += 1
    if created:
        db.session.commit()
    return created


def simulate_occupancy(percent: int, seed: int | None = 42) -> dict:
    """Setzt eine simulierte Grundbelegung pro Klinik (0..100 Prozent).

    Variiert per SK-Stufe mit ±10 % Jitter — realistischer als pauschal.
    """
    if not 0 <= percent <= 100:
        raise ValueError("percent muss 0..100 sein")
    init_belegung_rows()
    rng = random.Random(seed) if seed is not None else random.Random()

    rows = db.session.query(KrankenhausBelegung).all()
    for row in rows:
        row.vorbelegung_prozent = percent
        for sk in ("sk1", "sk2", "sk3"):
            cap = getattr(row, f"kapazitaet_{sk}")
            # Jitter: percent ±10, clamped 0..100
            jitter = rng.uniform(-10, 10)
            pct = max(0.0, min(100.0, percent + jitter))
            setattr(row, f"belegung_{sk}", int(round(cap * pct / 100.0)))
    db.session.commit()
    return {
        "updated": len(rows),
        "percent": percent,
    }


def reset_belegung() -> int:
    rows = db.session.query(KrankenhausBelegung).all()
    for row in rows:
        row.belegung_sk1 = 0
        row.belegung_sk2 = 0
        row.belegung_sk3 = 0
        row.vorbelegung_prozent = 0
    db.session.commit()
    return len(rows)


# ---------------- Dispatch ----------------

@dataclass
class DispatchResult:
    batch_id: int
    assigned: int
    unassigned: int
    transport_count: int
    avg_distanz_km: float | None


def _eligible_krankenhaeuser(sk: str):
    """Kliniken die einen Patienten der Stufe <sk> versorgen können.

    SK1 braucht kann_sk1; SK2 reicht kann_sk1 ODER kann_sk2;
    SK3 reicht kann_sk1 ODER kann_sk2 ODER kann_sk3.
    """
    cols = SK_KANN_COLS[sk]
    ors = [getattr(Krankenhaus, c) == True for c in cols]  # noqa: E712
    return (
        db.session.query(Krankenhaus, KrankenhausBelegung)
        .join(KrankenhausBelegung, KrankenhausBelegung.krankenhaus_id == Krankenhaus.id)
        .filter(Krankenhaus.lat.isnot(None), Krankenhaus.lon.isnot(None))
        .filter(Krankenhaus.ausgeschlossen == False)  # noqa: E712
        .filter(or_(*ors))
        .all()
    )


def dispatch_batch(batch: PatientenBatch, hub: Hub, use_here: bool = True) -> DispatchResult:
    """Verteilt alle pending-Patienten eines Batches auf Kliniken.

    Regeln:
    - Zuweisung in SK-Reihenfolge SK1 → SK2 → SK3 (vital zuerst)
    - Für jeden Patienten: nächste Klinik zuerst (Haversine-Vorranking zum Hub)
    - Nimmt erste Klinik mit freier Kapazität im benötigten SK-Segment
    - Radius erweitert sich automatisch weil wir nach Distanz sortieren
    - HERE Maps für echte Tourenplanung der ausgewählten Ziel-Klinik
    """
    init_belegung_rows()

    pending_patients = (
        batch.patients.filter_by(status="pending")
        .order_by(db.case((Patient.sk == "SK1", 1), (Patient.sk == "SK2", 2), else_=3),
                  Patient.eingangssichtung.asc().nulls_last())
        .all()
    )

    # Cache: pro SK-Stufe die eligiblen Kliniken mit Distanz (einmal berechnen)
    eligibles_by_sk: dict[str, list[tuple[Krankenhaus, KrankenhausBelegung, float]]] = {}
    for sk in ("SK1", "SK2", "SK3"):
        ranked = []
        for kh, bel in _eligible_krankenhaeuser(sk):
            d = haversine_km(hub.lat, hub.lon, kh.lat, kh.lon)
            ranked.append((kh, bel, d))
        ranked.sort(key=lambda t: t[2])
        eligibles_by_sk[sk] = ranked

    assigned = 0
    unassigned = 0
    transports: list[TransportAuftrag] = []
    distances: list[float] = []
    p_assignments: list[dict] = []

    for p in pending_patients:
        candidates = eligibles_by_sk.get(p.sk, [])
        target = None
        target_bel: KrankenhausBelegung | None = None
        target_dist = None
        for kh, bel, d in candidates:
            if bel.frei(p.sk) > 0:
                target = kh
                target_bel = bel
                target_dist = d
                break

        if target is None:
            p.status = "unassigned"
            p.note = f"Keine freie Kapazität für {p.sk} in ganz Deutschland."
            unassigned += 1
            continue

        # Patient zuweisen
        p.assigned_krankenhaus_id = target.id
        p.assigned_at = datetime.utcnow()
        p.aufenthaltsdauer_tage = SK_AUFENTHALT_TAGE[p.sk]
        p.distanz_km = target_dist
        p.status = "assigned"

        # Belegung hochzählen
        col = f"belegung_{p.sk.lower()}"
        setattr(target_bel, col, getattr(target_bel, col) + 1)

        # Zuweisung merken — Fahrt + Route bündeln wir nach dem Loop
        tm_info = IVENA_TRANSPORTMITTEL.get(p.sk, {})
        tm_code = tm_info.get("code", "KTW")
        p_assignments.append({
            "patient": p, "krankenhaus": target, "distance": target_dist, "tm_code": tm_code,
        })
        distances.append(target_dist)
        assigned += 1

    # ---------- Bündelung: (krankenhaus_id, transportmittel) → Fahrten ----------
    # Ziel: pro Fahrzeug genau 1 HERE-Abfrage, mehrere Patienten gebündelt.
    from collections import defaultdict
    buckets: dict[tuple[int, str], list[dict]] = defaultdict(list)
    for a in p_assignments:
        buckets[(a["krankenhaus"].id, a["tm_code"])].append(a)

    # Route-Cache je (kh_id) → HERE nur 1× pro Klinik, unabhängig vom Transportmittel
    route_cache: dict[int, dict] = {}
    fahrt_counter = 0
    fahrten_created = 0

    for (kh_id, tm_code), assigns in buckets.items():
        kapazitaet = TRANSPORTMITTEL_KAPAZITAET.get(tm_code, 1)
        target = assigns[0]["krankenhaus"]

        # Route holen — EIN HERE-Call pro Klinik, Ergebnis für alle Fahrten wiederverwenden
        if kh_id not in route_cache:
            if use_here:
                route = fetch_route((hub.lat, hub.lon), (target.lat, target.lon))
                import json as _json
                route_cache[kh_id] = {
                    "distanz_km": route.distance_km,
                    "dauer_min": route.duration_min,
                    "polyline": route.polyline,
                    "actions_json": (
                        None if not route.actions
                        else _json.dumps(route.actions, ensure_ascii=False)
                    ),
                    "geojson": (
                        None if not route.geojson else _json.dumps(route.geojson)
                    ),
                    "source": route.source,
                }
            else:
                # Haversine-Fallback
                d_km = haversine_km(hub.lat, hub.lon, target.lat, target.lon)
                route_cache[kh_id] = {
                    "distanz_km": round(d_km, 2),
                    "dauer_min": round(d_km / 70 * 60 * 1.3, 1),
                    "polyline": None,
                    "actions_json": None,
                    "geojson": None,
                    "source": "haversine",
                }
        rc = route_cache[kh_id]

        # SK1 zuerst innerhalb der Gruppe (vital zuerst ausfahren)
        assigns.sort(key=lambda a: {"SK1": 0, "SK2": 1, "SK3": 2}.get(a["patient"].sk, 3))

        # In Fahrten splitten nach Kapazität
        for i in range(0, len(assigns), kapazitaet):
            chunk = assigns[i:i + kapazitaet]
            fahrt_counter += 1
            fahrt_code = f"F-{batch.id:03d}-{fahrt_counter:04d}"

            # Abfahrt/Ankunft: frühester Transportbereit der Chunk-Patienten + Pickup-Offset
            ready_times = [a["patient"].transportbereit or a["patient"].eingangssichtung or datetime.utcnow() for a in chunk]
            base_time = max(ready_times)  # wir warten auf letzten im Bündel
            pickup_offset_min = TRANSPORTMITTEL_PICKUP_OFFSET_MIN.get(tm_code, 5)
            abfahrt_dt = base_time + timedelta(minutes=pickup_offset_min)
            ankunft_dt = abfahrt_dt + timedelta(minutes=rc["dauer_min"] or 0)

            fahrt = Fahrt(
                fahrt_code=fahrt_code,
                batch_id=batch.id,
                hub_id=hub.id,
                krankenhaus_id=kh_id,
                transportmittel=tm_code,
                kapazitaet=kapazitaet,
                anzahl_patienten=len(chunk),
                hub_lat=hub.lat, hub_lon=hub.lon,
                ziel_lat=target.lat, ziel_lon=target.lon,
                abfahrt=abfahrt_dt, ankunft=ankunft_dt,
                distanz_km=rc["distanz_km"], dauer_min=rc["dauer_min"],
                routing_source=rc["source"],
                route_geojson=rc["geojson"],
                here_polyline=rc["polyline"],
                here_instructions_json=rc["actions_json"],
            )
            db.session.add(fahrt)
            db.session.flush()
            fahrten_created += 1

            # Pro Patient einen TransportAuftrag mit fahrt_id + bundle_position
            for pos, a in enumerate(chunk, start=1):
                p = a["patient"]
                t = TransportAuftrag(
                    patient_id=p.id, batch_id=batch.id, fahrt_id=fahrt.id,
                    bundle_position=pos,
                    hub_id=hub.id, hub_lat=hub.lat, hub_lon=hub.lon,
                    krankenhaus_id=kh_id, ziel_lat=target.lat, ziel_lon=target.lon,
                    sk=p.sk,
                    distanz_km=rc["distanz_km"], dauer_min=rc["dauer_min"],
                    transportmittel=tm_code,
                    abfahrt=abfahrt_dt, ankunft=ankunft_dt,
                    here_polyline=rc["polyline"],
                    here_instructions_json=rc["actions_json"],
                    route_geojson=rc["geojson"],
                    routing_source=rc["source"],
                )
                db.session.add(t)
                transports.append(t)

    log.info("Dispatch: %d Patienten → %d Fahrten (HERE-Calls: %d)",
             assigned, fahrten_created, len(route_cache) if use_here else 0)

    batch.status = "dispatched"
    batch.dispatched_at = datetime.utcnow()
    db.session.commit()

    return DispatchResult(
        batch_id=batch.id,
        assigned=assigned,
        unassigned=unassigned,
        transport_count=len(transports),
        avg_distanz_km=round(sum(distances) / len(distances), 1) if distances else None,
    )


def reset_dispatch(batch: PatientenBatch) -> int:
    """Transportaufträge + Fahrten + Zuweisungen eines Batches zurücksetzen."""
    # Transportaufträge löschen
    deleted = (
        db.session.query(TransportAuftrag)
        .filter(TransportAuftrag.batch_id == batch.id)
        .delete()
    )
    db.session.query(Fahrt).filter(Fahrt.batch_id == batch.id).delete()
    # Belegungen reduzieren
    for p in batch.patients.filter(Patient.status == "assigned").all():
        if p.assigned_krankenhaus_id:
            bel = db.session.get(KrankenhausBelegung, p.assigned_krankenhaus_id)
            if bel and p.sk:
                col = f"belegung_{p.sk.lower()}"
                setattr(bel, col, max(getattr(bel, col) - 1, 0))
        p.status = "pending"
        p.assigned_krankenhaus_id = None
        p.assigned_at = None
        p.distanz_km = None
        p.note = None
    batch.status = "uploaded"
    batch.dispatched_at = None
    db.session.commit()
    return deleted
