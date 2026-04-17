"""XLSX-Simulator: generiert Patientenlisten im IVENA-ähnlichen Format.

Eingabe: Gesamtanzahl Verletzte, Verteilung SK1/SK2/SK3 (absolut oder Prozent),
          Zeitraum (Start-Datum + Anzahl Tage), Herkunft (z.B. Hub Süd).
Ausgabe: .xlsx-Datei in IVENA-Format + Spalte "Quelle" (Herkunft).
"""

from __future__ import annotations

import io
import random
from dataclasses import dataclass
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


COUNTRY_PREFIX = "DEU26"
LTU_PREFIX = "LTU"


@dataclass
class SimInput:
    sk1: int
    sk2: int
    sk3: int
    start_date: datetime
    days: int
    hub_name: str = "Hub Süd"
    start_hour: int = 8
    end_hour: int = 20
    seed: int | None = None

    @property
    def total(self) -> int:
        return self.sk1 + self.sk2 + self.sk3


def _random_time_between(start_dt: datetime, end_dt: datetime, rng: random.Random) -> datetime:
    span = int((end_dt - start_dt).total_seconds())
    offset = rng.randint(0, max(span, 1))
    return start_dt + timedelta(seconds=offset)


def generate_patients(inp: SimInput) -> list[dict]:
    """Liefert eine Liste von Patienten-Dicts (nicht noch als Excel)."""
    rng = random.Random(inp.seed) if inp.seed is not None else random.Random()

    sk_bucket: list[str] = (
        ["SK1"] * inp.sk1 +
        ["SK2"] * inp.sk2 +
        ["SK3"] * inp.sk3
    )
    rng.shuffle(sk_bucket)

    # Verteilung der Patienten über X Tage:
    # gleichmäßig aufteilen mit leichtem Jitter, Ankunft im Tagesfenster start_hour..end_hour
    days = max(inp.days, 1)
    per_day = len(sk_bucket) / days
    rows: list[dict] = []
    idx = 0
    for d in range(days):
        day_start = (inp.start_date + timedelta(days=d)).replace(hour=0, minute=0, second=0, microsecond=0)
        window_start = day_start.replace(hour=inp.start_hour)
        window_end = day_start.replace(hour=inp.end_hour)

        # Anzahl Patienten für diesen Tag: per_day + jitter, aber zusammen ergibt sich die Gesamtmenge
        count_today = int(round(per_day * (d + 1))) - idx
        count_today = max(count_today, 0)
        count_today = min(count_today, len(sk_bucket) - idx)

        for _ in range(count_today):
            sk = sk_bucket[idx]
            idx += 1
            sichtung = _random_time_between(window_start, window_end, rng)
            # Transportbereit: 2–20 min nach Sichtung
            transport = sichtung + timedelta(minutes=rng.randint(2, 20))
            patient_id = f"{COUNTRY_PREFIX}_{idx:03d}_{LTU_PREFIX}{idx:04d}"
            rows.append({
                "ID": patient_id,
                "SK": sk,
                "Datum": day_start.date(),
                "Eingangssichtung um": sichtung.time(),
                "Transportbereitschaft gesetzt um": transport.time(),
                "Quelle": inp.hub_name,
            })

    # Falls durch Rundung noch Patienten übrig: auf den letzten Tag packen
    while idx < len(sk_bucket):
        sk = sk_bucket[idx]
        idx += 1
        last_day_start = (inp.start_date + timedelta(days=days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)
        sichtung = _random_time_between(
            last_day_start.replace(hour=inp.start_hour),
            last_day_start.replace(hour=inp.end_hour),
            rng,
        )
        transport = sichtung + timedelta(minutes=rng.randint(2, 20))
        patient_id = f"{COUNTRY_PREFIX}_{idx:03d}_{LTU_PREFIX}{idx:04d}"
        rows.append({
            "ID": patient_id,
            "SK": sk,
            "Datum": last_day_start.date(),
            "Eingangssichtung um": sichtung.time(),
            "Transportbereitschaft gesetzt um": transport.time(),
            "Quelle": inp.hub_name,
        })

    rows.sort(key=lambda r: (r["Datum"], r["Eingangssichtung um"]))
    return rows


def build_xlsx(inp: SimInput) -> bytes:
    rows = generate_patients(inp)
    wb = Workbook()
    ws = wb.active
    ws.title = "Patienten"

    headers = ["ID", "SK", "Datum", "Eingangssichtung um", "Transportbereitschaft gesetzt um", "Quelle"]
    sk_colors = {"SK1": "FFB71C1C", "SK2": "FFF57C00", "SK3": "FFFBC02D"}

    # Header-Stil
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF343A40")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Zeilen
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r[key])
            if key == "SK":
                cell.fill = PatternFill("solid", fgColor=sk_colors[r["SK"]])
                cell.font = Font(bold=True, color="FFFFFFFF" if r["SK"] != "SK3" else "FF000000")
                cell.alignment = Alignment(horizontal="center")
            if key in ("Eingangssichtung um", "Transportbereitschaft gesetzt um"):
                cell.number_format = "HH:MM:SS"
            if key == "Datum":
                cell.number_format = "YYYY-MM-DD"

    # Summary-Sheet
    ws2 = wb.create_sheet("Zusammenfassung")
    ws2["A1"] = "MANV-Simulation"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Herkunft"
    ws2["B3"] = inp.hub_name
    ws2["A4"] = "Zeitraum"
    ws2["B4"] = f"{inp.start_date.date()} bis {(inp.start_date + timedelta(days=inp.days - 1)).date()} ({inp.days} Tage)"
    ws2["A5"] = "Gesamt"; ws2["B5"] = inp.total
    ws2["A6"] = "SK1"; ws2["B6"] = inp.sk1
    ws2["A7"] = "SK2"; ws2["B7"] = inp.sk2
    ws2["A8"] = "SK3"; ws2["B8"] = inp.sk3
    ws2["A10"] = "Generiert"; ws2["B10"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Spaltenbreiten
    widths = {"A": 26, "B": 8, "C": 12, "D": 22, "E": 34, "F": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 42

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
